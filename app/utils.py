# utils.py
import requests, os, json
from datetime import datetime
from kivy.utils import platform
import webbrowser

# CNPJ lookup (optional): usa BrasilAPI (padrão). Se houver limitação, a função retorna None.
CNPJ_API = "https://brasilapi.com.br/api/cnpj/v1/{}"

def consulta_cnpj_online(cnpj_clean):
    try:
        r = requests.get(CNPJ_API.format(cnpj_clean), timeout=6)
        if r.status_code == 200:
            data = r.json()
            # mapear campos para nosso app
            mapped = {
                "razao_social": data.get("razaoSocial"),
                "nome_fantasia": data.get("nomeFantasia"),
                "logradouro": data.get("logradouro"),
                "numero": data.get("numero"),
                "bairro": data.get("bairro"),
                "cep": data.get("cep"),
                "municipio": data.get("municipio"),
                "uf": data.get("uf"),
            }
            return mapped
    except Exception:
        return None
    return None

def format_phone_for_whatsapp(number):
    number = ''.join(ch for ch in number if ch.isdigit())
    if not number: return
    # whatsapp wa.me link
    url = f"https://wa.me/{number}"
    try:
        if platform == "android":
            import android
        webbrowser.open(url)
    except Exception:
        webbrowser.open(url)

# Export to Excel (xlsx) using openpyxl
def export_to_excel(rows):
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Clientes"
        headers = ["cnpj","nome","fantasia","telefone1","whatsapp","email","cidade","uf","valor_mensalidade","tipo_pagamento","tipo_contrato","observacoes","data_cadastro","teamviewer","anydesk"]
        ws.append(headers)
        for r in rows:
            tv = json.dumps(r.get("teamviewer",[]), ensure_ascii=False)
            ad = json.dumps(r.get("anydesk",[]), ensure_ascii=False)
            ws.append([r.get(k,"") for k in ["cnpj","nome","fantasia","telefone1","whatsapp","email","cidade","uf","valor_mensalidade","tipo_pagamento","tipo_contrato","observacoes","data_cadastro"]]+[tv,ad])
        path = os.path.join(os.getcwd(), f"clientes_export_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}.xlsx")
        wb.save(path)
        return path
    except Exception as e:
        return str(e)

def import_from_excel():
    try:
        from openpyxl import load_workbook
        fname = input("Caminho do arquivo xlsx para importar: ")
        if not fname or not os.path.exists(fname):
            return None, "Arquivo não encontrado"
        wb = load_workbook(fname)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        headers = rows[0]
        data_rows = rows[1:]
        db = __import__("db").db.Database()
        for r in data_rows:
            d = dict(zip(headers, r))
            # normalize minimal
            cliente = {
                "cnpj": d.get("cnpj",""),
                "nome": d.get("nome",""),
                "fantasia": d.get("fantasia",""),
                "telefone1": d.get("telefone1",""),
                "whatsapp": d.get("whatsapp",""),
                "email": d.get("email",""),
                "cidade": d.get("cidade",""),
                "uf": d.get("uf",""),
                "valor_mensalidade": d.get("valor_mensalidade",""),
                "tipo_pagamento": d.get("tipo_pagamento",""),
                "tipo_contrato": d.get("tipo_contrato",""),
                "observacoes": d.get("observacoes",""),
                "data_cadastro": d.get("data_cadastro",""),
                "teamviewer": json.loads(d.get("teamviewer") or "[]"),
                "anydesk": json.loads(d.get("anydesk") or "[]"),
            }
            db.insert_cliente(cliente)
        return fname, "OK"
    except Exception as e:
        return None, str(e)
